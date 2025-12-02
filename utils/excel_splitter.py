# utils/excel_splitter.py   geliştirilmiş RAM tüketmez

import asyncio
from pathlib import Path
from typing import Dict, List, Any, Optional

import xlsxwriter
from openpyxl import load_workbook

from utils.group_manager import group_manager
from utils.file_namer import generate_output_filename
from utils.logger import logger
from config import config


class ExcelSplitter:
    """
    Memory-safe, high-performance Excel splitter.
    - Uses XlsxWriter constant_memory mode (true streaming)
    - Extremely low RAM usage
    - Async group lookup
    """

    def __init__(self, input_path: str, headers: List[str]):
        self.input_path = input_path
        self.headers = headers

        # Runtime structures (very small in RAM)
        self.writers: Dict[str, xlsxwriter.Workbook] = {}
        self.sheets: Dict[str, Any] = {}
        self.row_counts: Dict[str, int] = {}
        self.matched_rows = 0  # İstatistik için
        self.unmatched_cities = set()  # Eşleşmeyen şehirler

    # ---------------------------------------------------------
    # Workbook and sheet creation
    # ---------------------------------------------------------
    async def _ensure_group_writer(self, group_id: str) -> None:
        """Create workbook + sheet for group if not exists."""
        if group_id in self.writers:
            return

        # Get group info (async)
        group_info = await group_manager.get_group_info(group_id)
        filename = await generate_output_filename(group_info)

        output_dir = config.paths.OUTPUT_DIR
        output_dir.mkdir(parents=True, exist_ok=True)

        file_path = output_dir / filename

        # Create streaming workbook
        wb = xlsxwriter.Workbook(
            file_path,
            {'constant_memory': True}   # KEY: streaming, low-RAM
        )
        ws = wb.add_worksheet("Veriler")

        # Write headers
        ws.write_row(0, 0, self.headers)

        self.writers[group_id] = wb
        self.sheets[group_id] = ws
        self.row_counts[group_id] = 1   # Next row index

        logger.debug(f"Writer created for group {group_id}: {file_path}")

    # ---------------------------------------------------------
    # Process one row
    # ---------------------------------------------------------
    async def _process_row(self, row: tuple) -> None:
        """Eşleşmeyen şehirleri de takip et"""
        city = row[1] if len(row) > 1 else None
        groups = await group_manager.get_groups_for_city(city)
        
        if not groups and city:
            self.unmatched_cities.add(city)
        
        for g in groups:
            await self._ensure_group_writer(g)
            ws = self.sheets[g]
            row_index = self.row_counts[g]
            ws.write_row(row_index, 0, row)
            self.row_counts[g] += 1
            self.matched_rows += 1


    # ---------------------------------------------------------
    # Main streaming executor
    # ---------------------------------------------------------
    async def run(self) -> Dict[str, Any]:
        try:
            logger.info("🔄 Group manager initializing…")
            await group_manager._ensure_initialized()

            logger.info("📥 Reading input file…")
            wb = load_workbook(self.input_path, read_only=True)
            ws = wb.active

            processed_rows = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                await self._process_row(row)
                processed_rows += 1

            wb.close()

            logger.info(f"✔ Processing complete. Total rows processed: {processed_rows}")

            return await self._finalize(processed_rows)

        except Exception as e:
            logger.error(f"❌ Error: {e}", exc_info=True)
            return {
                "success": False,
                "error": str(e),
                "total_rows": 0,
                "output_files": {},
            }

    # ---------------------------------------------------------
    # Save & close writers
    # ---------------------------------------------------------
    async def _finalize(self, processed_rows: int) -> Dict[str, Any]:
        output_files = {}

        for group_id, wb in self.writers.items():
            try:
                wb.close()
                ws_path = Path(wb.filename)

                output_files[group_id] = {
                    "filename": ws_path.name,
                    "path": ws_path,
                    "row_count": self.row_counts[group_id] - 1,
                }

                logger.info(f"📄 Saved: {ws_path.name}")

            except Exception as e:
                logger.error(f"Error closing workbook for {group_id}: {e}")

        return {
            "success": True,
            "total_rows": processed_rows,
            "matched_rows": self.matched_rows,  # ✅ EKLENDİ
            "output_files": output_files,
            "unmatched_cities": list(self.unmatched_cities),  # ✅ EKLENDİ
            "stats": self.row_counts.copy()  # ✅ EKLENDİ (veya başka bir stat)
        }
        

# ---------------------------------------------------------
# EXTERNAL API
# ---------------------------------------------------------
# ASYNC arayüz fonksiyonu
#async def split_excel_by_groups_streaming
async def split_excel_by_groups(input_path: str, headers: List[str]) -> Dict[str, Any]:
    splitter = ExcelSplitter(input_path, headers)
    return await splitter.run()

# SYNC arayüz (backward compatibility)
# def split_excel_by_groups_streaming_sync
def split_excel_by_groups_sync(input_path: str, headers: List[str]) -> Dict[str, Any]:
    """Sync wrapper."""
    return asyncio.run(split_excel_by_groups(input_path, headers))
